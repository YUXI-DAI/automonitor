import traceback

from openpyxl import Workbook, load_workbook


class devData:
    sendfree = False
    sendtop=False
    sendinvalid=False
    receivefree = []
    receivetop=[]
    devicename=""
    data=[]

    def combine(self):
        count = 0
        for i in self.receivefree:
            j=self.receivetop[count]
            self.data.append((i[0],i[1],j[1]))
            count=count+1
    def createxel(self):
        try:
            wb = Workbook()
            sheet = wb.active
            sheet['A1'] = "设备"
            sheet['B1'] = "内存情况"
            sheet['C1'] = "CPU情况"
            wb.save("test.xlsx")
        except:
            raise Exception("excel文件创建失败")

    def addtoexcel(self):
        try:
            wb = load_workbook('test.xlsx')
            sheet = wb.active
            for i in self.data:
                sheet.append(i)
            wb.save('test.xlsx')
        except FileNotFoundError:
            raise Exception("提示：数据添加失败，找不到该excel文件")
        except PermissionError:
            raise Exception("提示：数据添加失败，该excel已经被打开，请关闭")
        except:
            traceback.print_exc()
            raise Exception("提示：数据添加失败，请排查原因")



