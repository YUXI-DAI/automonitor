import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import re
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# wb=Workbook()
#
# wb.save("test.xlsx")
wb = Workbook()
sheet = wb.active
sheet['A1'] = "设备"
sheet['B1'] = "内存情况"
sheet['C1'] = "CPU情况"
wb.save("test.xlsx")

wb = load_workbook('test.xlsx')
sheet = wb.active

# ('192.168.1.8', 'free\r\n             total         used         free       shared      buffers\r\nMem:        253100        82268       170832         4268            0\r\n-/+ buffers:              82268       170832\r\nSwap:            0            0            0\r\n~ # '),
text= 'top\r\n\x1b[H\x1b[JMem: 82280K used, 170820K free, 0K shrd, 0K buff, 39332K cached\r\nCPU:  0.0% usr 54.5% sys  0.0% nic 45.4% idle  0.0% io  0.0% irq  0.0% sirq\r\nLoad average: 19.00 17.15 16.93 2/194 13044\r\n\x1b[7m  PID  PPID USER     STAT   VSZ %VSZ CPU %CPU COMMAND\x1b[0m\r\n  330   326 root     S    66312 26.1   1 45.4 app_hdmi_io.bin\r\n  346   343 root     S     143m 57.8   0  0.0 gbusbdisplay\r\n  340   338 root     S     132m 53.3   0  0.0 gbairplay\r\n  386   382 root     S    83892 33.0   1  0.0 miracast\r\n  358   349 root     S    56244 22.1   0  0.0 gbcast\r\n  356   353 root     S    53276 20.9   0  0.0 miniosd\r\n  364   361 root     S    45716 18.0   1  0.0 fsc600media\r\n  324   321 root     S    31996 12.6   0  0.0 devmanageserver\r\n  612   605 root     S    27284 10.7   0  0.0 mainserver\r\n  390   387 root     S    21092  8.3   1  0.0 gbnetfilter\r\n  636     1 root     S    17804  7.0   0  0.0 /usr/local/bin/filewatch -s /confi\r\n  915   607 root     S     8796  3.4   0  0.0 web\r\n  627   623 root     S     6432  2.5   1  0.0 hardwatchdog hardwatchdog app_led_\r\n  623     1 root     S     4936  1.9   0  0.0 gbcmd hardwatchdog hardwatchdog ap\r\n  343     1 root     S     4936  1.9   1  0.0 gbcmd gbusbdisplay\r\n  326     1 root     S     4936  1.9   0  0.0 gbcmd app_hdmi_io.bin\r\n  605     1 root     S     4936  1.9   0  0.0 gbcmd mainserver\r\n  312     1 root     S     4936  1.9   1  0.0 gbcmd app_led_ctrl.bin\r\n  349     1 root     S     4936  1.9   1  0.0 gbcmd gbcast\r\n  607     1 root     S     4936  1.9   1  0.0 gbcmd web\r'
text= 'top\r\n[H[JMem: 82280K used, 170820K free, 0K shrd, 0K buff, 39332K cached\r\nCPU:  0.0% usr 54.5% sys  0.0% nic 45.4% idle  0.0% io  0.0% irq  0.0% sirq\r\nLoad average: 19.00 17.15 16.93 2/194 13044\r\n[7m  PID  PPID USER     STAT   VSZ %VSZ CPU %CPU COMMAND[0m\r\n  330   326 root     S    66312 26.1   1 45.4 app_hdmi_io.bin\r\n  346   343 root     S     143m 57.8   0  0.0 gbusbdisplay\r\n  340   338 root     S     132m 53.3   0  0.0 gbairplay\r\n  386   382 root     S    83892 33.0   1  0.0 miracast\r\n  358   349 root     S    56244 22.1   0  0.0 gbcast\r\n  356   353 root     S    53276 20.9   0  0.0 miniosd\r\n  364   361 root     S    45716 18.0   1  0.0 fsc600media\r\n  324   321 root     S    31996 12.6   0  0.0 devmanageserver\r\n  612   605 root     S    27284 10.7   0  0.0 mainserver\r\n  390   387 root     S    21092  8.3   1  0.0 gbnetfilter\r\n  636     1 root     S    17804  7.0   0  0.0 /usr/local/bin/filewatch -s /confi\r\n  915   607 root     S     8796  3.4   0  0.0 web\r\n  627   623 root     S     6432  2.5   1  0.0 hardwatchdog hardwatchdog app_led_\r\n  623     1 root     S     4936  1.9   0  0.0 gbcmd hardwatchdog hardwatchdog ap\r\n  343     1 root     S     4936  1.9   1  0.0 gbcmd gbusbdisplay\r\n  326     1 root     S     4936  1.9   0  0.0 gbcmd app_hdmi_io.bin\r\n  605     1 root     S     4936  1.9   0  0.0 gbcmd mainserver\r\n  312     1 root     S     4936  1.9   1  0.0 gbcmd app_led_ctrl.bin\r\n  349     1 root     S     4936  1.9   1  0.0 gbcmd gbcast\r\n  607     1 root     S     4936  1.9   1  0.0 gbcmd web\r'
print(text)


text= ILLEGAL_CHARACTERS_RE.sub(r'', text)

print(text)
data=[]
data.append(('da',text))
print(data)
for i in data:
    sheet.append(i)
wb.save("test.xlsx")